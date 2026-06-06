"""测试用例输出模块 - Excel 和 Markdown"""

import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["用例编号", "模块", "标题", "前置条件", "测试步骤", "预期结果", "优先级", "用例类型"]


def _strip_trailing_punctuation(text: str) -> str:
    """去除文本结尾的标点符号"""
    if not text:
        return text
    text = text.rstrip()
    while text and text[-1] in ("。", ".", "，", ",", "；", ";", "：", ":"):
        text = text[:-1].rstrip()
    return text


def _normalize_steps(steps: str) -> str:
    """统一测试步骤格式：去除每步结尾的标点"""
    if not steps:
        return steps
    lines = steps.split("\n")
    normalized = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^(\d+[.、]\s*)', line)
        if m:
            prefix = m.group(1)
            content = line[len(prefix):].strip()
            while content and content[-1] in ("。", ".", "，", ",", "；", ";", "：", ":"):
                content = content[:-1].rstrip()
            normalized.append(prefix + content)
        else:
            normalized.append(line)
    return "\n".join(normalized)


def normalize_testcase(tc: dict) -> dict:
    """规范化用例文本字段，确保导出风格一致"""
    tc = tc.copy()
    tc["expected"] = _strip_trailing_punctuation(tc.get("expected", ""))
    tc["precondition"] = _strip_trailing_punctuation(tc.get("precondition", ""))
    tc["steps"] = _normalize_steps(tc.get("steps", ""))
    return tc


PRIORITY_COLORS = {
    "P0": "FF0000",  # 红
    "P1": "FF6600",  # 橙
    "P2": "FFCC00",  # 黄
    "P3": "00CC00",  # 绿
}


def _split_steps(steps: str) -> list[str]:
    """智能分割测试步骤，兼容换行和分号两种格式"""
    if "\n" in steps:
        return steps.split("\n")
    # 兼容分号分隔的格式：1. xxx; 2. xxx
    parts = re.split(r';\s*(?=\d+\.)', steps)
    return parts if len(parts) > 1 else [steps]


def _ensure_dir(path: str):
    Path(path).mkdir(parents=True, exist_ok=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def to_excel(testcases: list[dict], output_dir: str, filename: str | None = None) -> str:
    """导出为 Excel"""
    _ensure_dir(output_dir)
    fname = filename or f"testcases_{_timestamp()}.xlsx"
    filepath = os.path.join(output_dir, fname)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 数据行
    for row_idx, tc in enumerate(testcases, 2):
        tc = normalize_testcase(tc)
        values = [
            tc.get("id", ""),
            tc.get("module", ""),
            tc.get("title", ""),
            tc.get("precondition", ""),
            tc.get("steps", ""),
            tc.get("expected", ""),
            tc.get("priority", ""),
            tc.get("type", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # 优先级着色
            if col_idx == 7:
                color = PRIORITY_COLORS.get(str(val), "")
                if color:
                    cell.font = Font(bold=True, color=color)

    # 列宽
    widths = [12, 15, 30, 20, 40, 30, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 筛选
    ws.auto_filter.ref = f"A1:H{len(testcases) + 1}"

    wb.save(filepath)
    return filepath


def to_markdown(testcases: list[dict], output_dir: str, filename: str | None = None) -> str:
    """导出为 Markdown"""
    _ensure_dir(output_dir)
    fname = filename or f"testcases_{_timestamp()}.md"
    filepath = os.path.join(output_dir, fname)

    lines = [
        "# 测试用例",
        "",
        f"> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"> 用例总数: {len(testcases)}",
        "",
    ]

    # 按模块分组
    modules: dict[str, list[dict]] = {}
    for tc in testcases:
        mod = tc.get("module", "未分类")
        modules.setdefault(mod, []).append(tc)

    for mod, tcs in modules.items():
        lines.append(f"## {mod}")
        lines.append("")
        lines.append("| 编号 | 标题 | 优先级 | 类型 |")
        lines.append("|------|------|--------|------|")
        for tc in tcs:
            lines.append(
                f"| {tc.get('id', '')} | {tc.get('title', '')} "
                f"| {tc.get('priority', '')} | {tc.get('type', '')} |"
            )
        lines.append("")

        # 详细用例
        for tc in tcs:
            tc = normalize_testcase(tc)
            lines.append(f"### {tc.get('id', '')} - {tc.get('title', '')}")
            lines.append("")
            lines.append(f"- **模块**: {tc.get('module', '')}")
            lines.append(f"- **优先级**: {tc.get('priority', '')}")
            lines.append(f"- **类型**: {tc.get('type', '')}")
            if tc.get("precondition"):
                lines.append(f"- **前置条件**: {tc.get('precondition', '')}")
            lines.append(f"- **测试步骤**:")
            for step in _split_steps(tc.get("steps", "")):
                step = step.strip()
                if step:
                    lines.append(f"  {step}")
            lines.append(f"- **预期结果**: {tc.get('expected', '')}")
            lines.append("")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return filepath
