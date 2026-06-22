"""需求文档读取模块 - 支持 Markdown/TXT/Excel/图片/手动输入"""

import base64
import os
from pathlib import Path

from openpyxl import load_workbook

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}


def read_text(path: str) -> str:
    """读取 Markdown 或 TXT 文件"""
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def read_excel(path: str) -> str:
    """读取 Excel 需求文件，拼接为文本"""
    wb = load_workbook(path, read_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c else "" for c in rows[0]]
        parts.append(f"## Sheet: {ws.title}")
        parts.append(" | ".join(header))
        parts.append(" | ".join(["---"] * len(header)))
        for row in rows[1:]:
            cells = [str(c) if c else "" for c in row]
            parts.append(" | ".join(cells))
        parts.append("")
    wb.close()
    return "\n".join(parts)


def read_image(path: str) -> bytes:
    """读取图片文件，返回原始字节"""
    with open(path, "rb") as f:
        return f.read()


def image_to_base64(path: str) -> str:
    """将图片转为 base64 字符串"""
    data = read_image(path)
    return base64.b64encode(data).decode("utf-8")


def get_image_media_type(path: str) -> str:
    """根据文件扩展名返回 MIME 类型"""
    ext = Path(path).suffix.lower()
    mime_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
    }
    return mime_map.get(ext, "image/png")


def is_image(path: str) -> bool:
    """判断文件是否为图片"""
    return Path(path).suffix.lower() in IMAGE_EXTS


def read_manual() -> str:
    """手动输入需求"""
    print("\n请输入需求描述（输入空行结束）：")
    lines = []
    while True:
        try:
            line = input()
            if line.strip() == "":
                if lines:
                    break
                continue
            lines.append(line)
        except EOFError:
            break
    return "\n".join(lines)


def read_requirement(source: str | None) -> str:
    """统一入口：根据来源类型读取纯文本需求（命令行用）"""
    if source is None:
        return read_manual()

    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {source}")

    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        return read_excel(source)
    elif ext in IMAGE_EXTS:
        raise ValueError(f"图片文件请使用 Web 版上传，或通过 --image 参数指定")
    else:
        return read_text(source)


def read_requirement_multimodal(source: str | None, image_paths: list[str] | None = None) -> dict:
    """多模态入口：返回 {text: str, images: [{data, media_type}]}"""
    result = {"text": "", "images": []}

    # 读取文本（source 为图片路径时跳过文本读取）
    if source and not is_image(source):
        result["text"] = read_requirement(source)
    # source 为 None 时：如果有图片则跳过（Web 场景），否则 fallback 到手动输入
    elif source is None and not image_paths:
        result["text"] = read_manual()

    # 读取图片
    if image_paths:
        for p in image_paths:
            if not os.path.exists(p):
                raise FileNotFoundError(f"图片文件不存在: {p}")
            result["images"].append({
                "data": image_to_base64(p),
                "media_type": get_image_media_type(p),
                "filename": Path(p).name,
            })

    return result
